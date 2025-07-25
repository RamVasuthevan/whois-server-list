import csv
import os
import time
from dataclasses import dataclass
from typing import List, Optional, Union
from zipfile import ZIP_DEFLATED, ZipFile

import requests
from tenacity import retry, stop_after_attempt, wait_exponential
from bs4 import BeautifulSoup
from openpyxl import Workbook as OriginalWorkbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook


def save_workbook(workbook, filename):
    """Save the given workbook on the filesystem under the name filename.

    :param workbook: the workbook to save
    :type workbook: :class:`openpyxl.workbook.Workbook`

    :param filename: the path to which save the workbook
    :type filename: string

    :rtype: bool
    """
    archive = ZipFile(filename, "w", ZIP_DEFLATED, allowZip64=True)
    # workbook.properties.modified = datetime.datetime.utcnow()
    writer = ExcelWriter(workbook, archive)
    writer.save()
    return True


class Workbook(OriginalWorkbook):
    def save(self, filename: str) -> None:
        """Save the current workbook under the given `filename`.
        Use this function instead of using an `ExcelWriter`.

        .. warning::
            When creating your workbook using `write_only` set to True,
            you will only be able to call this function once. Subsequent attempts to
            modify or save the file will raise an :class:`openpyxl.shared.exc.WorkbookAlreadySaved` exception.
        """
        if self.read_only:
            raise TypeError("""Workbook is read-only""")
        if self.write_only and not self.worksheets:
            self.create_sheet()
        save_workbook(self, filename)


@dataclass
class Result:
    tld_punycode: str
    whois_server_url: str
    tld_unicode: Optional[Union[None, str]] = None


def create_csv(results: List[Result]):
    HEADERS = ["Domain", "WHOIS Server URL"]
    FILENAME = os.path.join(os.pardir, "whois-servers.csv")

    with open(FILENAME, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(HEADERS)
        writer.writerows(
            ((result.tld_punycode, result.whois_server_url) for result in results)
        )


def create_markdown(results: List[Result]):
    FILENAME = os.path.join(os.pardir, "whois-servers.md")

    with open(FILENAME, "w", newline="", encoding="utf-8") as file:
        file.write("| Domain   | WHOIS Server URL          |\n")
        file.write("|----------|--------------------------|\n")
        for result in results:
            file.write(f"| {result.tld_punycode} | {result.whois_server_url} |\n")


def create_xlsx(results: List[Result]):
    HEADERS = ["Domain", "WHOIS Server URL"]
    FILENAME = os.path.join(os.pardir, "whois-servers.xlsx")

    wb = Workbook()
    ws = wb.active

    ws.append(HEADERS)

    for result in results:
        ws.append([result.tld_punycode, result.whois_server_url])

    old_wb = load_workbook(FILENAME)
    wb.properties = old_wb.properties

    wb.save(FILENAME)


def create_README(results: List[Result]):
    BEFORE_FILENAME = "README/before.md"
    AFTER_FILENAME = "README/after.md"
    README_FILENAME = os.path.join(os.pardir, "README.md")

    with open(BEFORE_FILENAME, "r", encoding="utf-8") as before_file:
        before_content = before_file.read()

    markdown_content = "| Domain   | WHOIS Server URL          |\n"
    markdown_content += "|----------|--------------------------|\n"
    for result in results:
        if result.tld_punycode == result.tld_unicode.lower():
            markdown_content += f"| .{result.tld_punycode} | [{result.whois_server_url}](https://{result.whois_server_url})|\n"
        else:
            markdown_content += f"| .{result.tld_punycode} (.{result.tld_unicode}) | [{result.whois_server_url}](https://{result.whois_server_url}) |\n"

    with open(AFTER_FILENAME, "r", encoding="utf-8") as after_file:
        after_content = after_file.read()

    readme_content = before_content + markdown_content + after_content

    with open(README_FILENAME, "w", encoding="utf-8") as readme_file:
        readme_file.write(readme_content)


if __name__ == "__main__":
    MAX_RETRIES = 3
    TIMEOUT = 10
    SLEEP = 1
    MIN_WAIT = 4
    MAX_WAIT = 10

    BASE_IANA_URL = "https://www.iana.org"
    ROOT_ZONE_DATABASE_URL = BASE_IANA_URL + "/domains/root/db"

    @retry(
        stop=stop_after_attempt(MAX_RETRIES),
        wait=wait_exponential(multiplier=1, min=MIN_WAIT, max=MAX_WAIT),
    )
    def get_response(url):
        response = requests.get(url, timeout=TIMEOUT)
        response.raise_for_status()
        return response

    response = get_response(ROOT_ZONE_DATABASE_URL)

    if response and response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        table = soup.find("table", id="tld-table")
        results = []

        for tr in table.tbody.find_all("tr"):
            relative_tld_url = tr.td.span.a.get("href")
            tld_punycode = os.path.splitext(os.path.basename(relative_tld_url))[0]
            print(tld_punycode)

            response = get_response(BASE_IANA_URL + relative_tld_url)

            if response and response.status_code == 200:
                soup = BeautifulSoup(response.content, "html.parser")
                tld_unicode = soup.find("h1").text.split(".")[-1].strip()
                registry_information = soup.find(
                    "h2", string="Registry Information"
                ).find_next("p")

                if registry_information.find("b", string="WHOIS Server:"):
                    whois_url = registry_information.find(
                        "b", string="WHOIS Server:"
                    ).next_sibling.strip()
                    results.append(Result(tld_punycode, whois_url, tld_unicode))
            else:
                print(
                    f"Error occurred while fetching {tld_punycode}. Status Code: {response.status_code}"
                )
                response.raise_for_status()

            time.sleep(SLEEP)

if __name__ == "__main__":
    create_csv(results)
    create_markdown(results)
    create_README(results)
    create_xlsx(results)
